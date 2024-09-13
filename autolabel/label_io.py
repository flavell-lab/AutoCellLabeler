import openpyxl, csv


def export_excel_to_csv(file_path, output_csv_path, dataset):
    """
    Extracts data from a specific worksheet in an Excel file and exports it to a CSV file.

    This function opens an Excel workbook, filters the sheets based on the presence of the specified 
    `dataset` name or the keyword "labels" in their names while excluding sheets with "progress" 
    and "alt" in their names. It selects the last sheet from the filtered list, extracts its data, 
    and writes this data to a CSV file.

    Args:
        file_path (str): The file path of the input Excel workbook.
        output_csv_path (str): The file path where the output CSV will be saved.
        dataset (str): A keyword to filter the sheet names within the Excel workbook.

    Returns:
        None
    """
    # Open the Excel workbook
    wb = openpyxl.load_workbook(file_path)
    
    # Get all sheet names
    sheet_names = wb.sheetnames
    
    # Filter the sheet names
    filtered_sheet_names = [name for name in sheet_names if ("labels" in name or dataset in name) and "progress" not in name and "alt" not in name]
    
    # Select the last sheet name from the filtered list
    selected_sheet_name = filtered_sheet_names[-1]
    sheet = wb[selected_sheet_name]
    
    # Extract data from the selected sheet
    data = [[cell for cell in row] for row in sheet.iter_rows(values_only=True)]
    
    # Export the data to a CSV file
    with open(output_csv_path, 'w', newline='') as csvfile:
        writer = csv.writer(csvfile)
        writer.writerows(data)



